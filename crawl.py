import requests as rq
from bs4 import BeautifulSoup as bs
import time,os,re
from config import headers
from openpyxl import Workbook

class Coupang:
    @staticmethod
    def get_product_code(url: str)-> str:
        """ 입력받은 URL 주소의 PRODUCT CODE 추출하는 메소드 """
        prod_code = url.split('products/')[-1].split('?')[0]
        return prod_code

    def __init__(self):
        self.__headers = headers.get_headers(key='headers')

    def main(self)-> list:
        # URL 주소
        URL = self.input_review_url()

        # URL의 Product Code 추출
        prod_code = self.get_product_code(url=URL)

        # URL 주소 재가공
        URLS = [f'https://www.coupang.com/vp/product/reviews?productId={prod_code}&page={page}&size=5&sortBy=ORDER_SCORE_ASC&ratings=&q=&viRoleCode=3&ratingSummary=true' for page in range(1,self.input_page_count() + 1)]

        # __headers에 referer 키 추가
        self.__headers['referer'] = URL

        with rq.Session() as session:
            return [self.fetch(url=url,session=session) for url in URLS]

    def fetch(self,url:str,session)-> list:
        save_data = list()

        with session.get(url=url,headers=self.__headers) as response :
            html = response.text
            soup = bs(html,'html.parser')

            # Article Boxes
            article_lenth = len(soup.select('article.sdp-review__article__list'))

            for idx in range(article_lenth):
                dict_data = dict()
                articles = soup.select('article.sdp-review__article__list')

                # 구매자 이름
                user_name = articles[idx].select_one('span.sdp-review__article__list__info__user__name')
                if user_name == None or user_name.text == '':
                    user_name = '-'
                else:
                    user_name = user_name.text.strip()

                # 평점
                rating = articles[idx].select_one('div.sdp-review__article__list__info__product-info__star-orange')
                if rating == None:
                    rating = 0
                else :
                    rating = int(rating.attrs['data-rating'])

                # 구매자 상품명
                prod_name = articles[idx].select_one('div.sdp-review__article__list__info__product-info__name')
                if prod_name == None or prod_name.text == '':
                    prod_name = '-'
                else:
                    prod_name = prod_name.text.strip()

                # 헤드라인(타이틀)
                headline = articles[idx].select_one('div.sdp-review__article__list__headline')
                if headline == None or headline.text == '':
                    headline = '등록된 헤드라인이 없습니다'
                else:
                    headline = headline.text.strip()

                # 리뷰 내용
                review_content = articles[idx].select_one('div.sdp-review__article__list__review > div')
                if review_content == None :
                    review_content = '등록된 리뷰내용이 없습니다'
                else:
                    review_content = re.sub('[\n\t]','',review_content.text.strip())

                # 도움 수
                answer = articles[idx].select_one('span.js_reviewArticleHelpfulCount')
                if answer == None or answer.text == '이 상품평이 도움 되었나요?':
                    answer = '0명에게 도움 됨'
                else:
                    answer = answer.text.strip() + "명에게 도움 됨"

                dict_data['상품명'] = prod_name
                dict_data['구매자 ID'] = user_name
                dict_data['구매자 별점'] = rating
                dict_data['리뷰 제목'] = headline
                dict_data['리뷰 내용'] = review_content
                dict_data['리뷰 도움 수'] = answer

                save_data.append(dict_data)

                print(dict_data , '\n')

            time.sleep(1)

            return save_data

    def input_review_url(self)-> str:
        while True:
            os.system('cls')
            review_url = input('원하시는 상품의 URL 주소를 입력해주세요\n\nEx)\nhttps://www.coupang.com/vp/products/6451503812?itemId=14007944553&vendorItemId=73528488680&sourceType=srp_product_ads&clickEventId=28aaab30-71e3-4f30-9059-07a29eb1b27f&korePlacement=15&koreSubPlacement=6&q=%EB%9E%A9%EB%85%B8%EC%89%AC&itemsCount=36&searchId=af6bda06076947a39f847ed86a718c34&rank=5&isAddedCart=\n\n:')
            if not review_url :
                os.system('cls')
                print('URL 주소가 입력되지 않았습니다')
                continue
            return review_url

    def input_page_count(self)-> int:
        os.system('cls')
        while True:
            page_count = input('페이지 수를 입력하세요\n\n:')
            if not page_count:
                print('페이지 수가 입력되지 않았습니다\n')
                continue

            return int(page_count)

class OpenPyXL:
    @staticmethod
    def save_file():
        # 크롤링 결과
        results : list = Coupang().main()

        wb = Workbook()
        ws = wb.active
        ws.append(['상품명','구매자 ID','구매자 별점','리뷰 제목','리뷰 내용','도움 수'])

        row = 2

        for x in results:
            for result in x :
                ws[f'A{row}'] = result['상품명']
                ws[f'B{row}'] = result['구매자 ID']
                ws[f'C{row}'] = result['구매자 별점']
                ws[f'D{row}'] = result['리뷰 제목']
                ws[f'E{row}'] = result['리뷰 내용']
                ws[f'F{row}'] = result['리뷰 도움 수']

                row += 1

        savePath = os.path.abspath('쿠팡 상품리뷰 크롤링')
        fileName = results[0][0]['상품명'] + '.xlsx'

        if not os.path.exists(savePath):
            os.mkdir(savePath)

        wb.save(os.path.join(savePath,fileName))
        wb.close()

        print(f'파일 저장완료!\n\n{os.path.join(savePath,fileName)}')
